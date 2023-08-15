import pandas as pd
import os
import warnings
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


class Growth:
    def __init__(self, path_file, path_seed):
        self.path_file = path_file
        self.path_seed = path_seed
        self.df_seed = self.get_df()
        self.room_name = self.df_seed.columns[-1]
        self.file_name = self.room_name + '端口占用表.xlsx'
        self.list_rack = list({}.fromkeys(self.df_seed['机架编号'].to_list()).keys())
        self.list_rack_name = [item + '端截面图' for item in self.list_rack]
        self.dict_areas = self.calculate_areas()

    def get_df(self):
        df = pd.read_excel(self.path_file + self.path_seed)
        df = df.astype(str)
        return df

    def calculate_areas(self):
        dict_rack = {}
        for i in self.list_rack:
            df_t = self.df_seed[self.df_seed['机架编号'] == i]
            df_t['quantity'] = df_t.apply(lambda x: self.calculate_areas_fun1(x['端口']), axis=1)
            dict_rack[i] = sum(df_t['quantity']) / 12
        return dict_rack

    def calculate_areas_fun1(self, str_):
        list_t = str_.split('-')
        return int(list_t[1]) - int(list_t[0]) + 1

    def generate_sheets(self):
        list_ = self.list_rack_name
        wb = Workbook()
        ws = wb.active
        ws.title = str(list_[0])
        for i in list_[1:]:
            wb.create_sheet(str(i))
        wb.save(self.file_name)

    def draw_border(self, ws, int_):
        str_ = 'A1:M' + str(int_)
        area = ws[str_]
        for i in area:
            for j in i:
                j.border = Border(right=Side(style='thin'), bottom=Side(style='thin'))

    def fill_color(self, ws, int_):
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
        for i in range(2, 14):
            for j in [int_ * 3, int_ * 3 + 1]:
                ws.cell(row=j, column=i, value="").fill = fill

    def write_frames(self):
        list_ = self.list_rack_name
        list_2 = self.list_rack
        str_ = self.file_name
        dict_ = self.dict_areas
        for i in range(len(list_)):
            wb = load_workbook(path_file + '/' + self.file_name)
            ws = wb[str(list_[i])]
            font = Font(u'宋体', size=11, bold=True, italic=False, strike=False, color='000000')
            ws.cell(row=1, column=1, value=str_ + '-' + list_2[i]).font = font
            ws.merge_cells(range_string='A1:M1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            quantity = dict_[list_2[i]]
            self.draw_border(ws, int(quantity * 3 + 1))
            for j in range(1, int(quantity + 1)):
                ws.cell(row=j * 3 - 1, column=1, value='标签').font = font
                ws.cell(row=j * 3, column=1, value='端子').font = font
                str_2 = 'A' + str(j * 3) + ':A' + str(j * 3 + 1)
                ws.merge_cells(range_string=str_2)
                str_3 = 'A' + str(j * 3)
                str_4 = 'A' + str(j * 3 - 1)
                ws[str_3].alignment = Alignment(horizontal='center', vertical='center')
                ws[str_4].alignment = Alignment(horizontal='left', vertical='center')
                self.fill_color(ws, j)
            wb.save(self.file_name)

    def write_details(self):
        list_ = self.list_rack_name
        list_2 = self.list_rack
        for i in range(len(list_)):
            df_t = self.df_seed[self.df_seed['机架编号'] == list_2[i]]
            df_t['from'] = df_t.apply(lambda x: int(x['端口'].split('-')[0]), axis=1)
            df_t['to'] = df_t.apply(lambda x: int(x['端口'].split('-')[1]), axis=1)
            wb = load_workbook(path_file + '/' + self.file_name)
            ws = wb[str(list_[i])]
            row_number = 2
            if len(df_t) == 1:
                series_ = df_t.iloc[0, :]
                n = self.write_details_fun1(ws, series_, row_number)
                row_number = n
            else:
                for j in range(len(df_t)):
                    series_ = df_t.iloc[j, :]
                    n = self.write_details_fun1(ws, series_, row_number)
                    row_number = n
            wb.save(self.file_name)

    def prune(self, str_):
        if str_[0:4] == 'ODF-':
            return str_[4:]
        elif str_ == 'ODF':
            return ''
        elif str_[0:3] == 'ODF':
            return str_[3:]
        else:
            return str_

    def write_details_fun1(self, ws, series_, int_):
        font = Font(u'Arial', size=11, bold=False, italic=False, strike=False, color='000000')
        align = Alignment(horizontal='center', vertical='center')
        a = [x for x in range(1, series_['to'] + 1)]
        list_ = [a[x:x + 12] for x in range(series_['from'] - 1, series_['to'], 12)]
        for i in list_:
            str_ = series_['机架编号'] + '-' + series_['子架编号'] + '/' + str(i[0]) + '-' + str(i[-1]) + '芯'
            ws.cell(row=int_, column=2, value=str_).font = font
            str_2 = 'B' + str(int_) + ':' + 'M' + str(int_)
            str_3 = 'B' + str(int_)
            ws.merge_cells(range_string=str_2)
            ws[str_3].alignment = Alignment(horizontal='center', vertical='center')
            for j in range(12):
                t = self.prune(series_['子架编号'])
                str_4 = t + '/' + str(i[j])
                ws.cell(row=int_ + 1, column=j + 2, value=str_4).font = font
                ws.cell(row=int_ + 1, column=j + 2, value=str_4).alignment = align
            int_ += 3
        return int_


if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    path_file = os.getcwd()
    path_seed = r'\seed.xlsx'
    gw = Growth(path_file=path_file, path_seed=path_seed)
    gw.generate_sheets()
    gw.write_frames()
    gw.write_details()
